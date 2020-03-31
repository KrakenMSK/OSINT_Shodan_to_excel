from openpyxl import load_workbook
import xlwt
import openpyxl
import re
import glob
import subprocess
import os
from shodan import Shodan
import ipaddress
import requests
import time
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import argparse


def parser():
    parser = argparse.ArgumentParser(description="calculate X to the power of Y")
    parser.add_argument("shodan_key", help="Key for shodan to collect info (32 symbols)")
    parser.add_argument("path_to_nets", help="Path to file with networks (available format: x.x.x.x of x.x.x.x. Every network - new string)")
    args = parser.parse_args()
    return args.shodan_key, args.path_to_nets

key, path_to_file = parser()
api = Shodan(key)

def checkNet():
    for i in range(5000):
        try:
            response = requests.get("http://www.google.com")
            if i!=0:
                print('Internet connection is ok')
            return True
        except:
            time.sleep(1)
            if i%5==0:
                print('There is not internet connection. Try to reset it. You have', str(5000-i), 'seconds to reset your internet connection')
            i+=1
    print('There is not Internet connection. Try to reset it and start programm')
    return False

def get_nets_from_file(IP_file):
    mass_all_info=''
    f=open(IP_file,'r')
    nets_to_research=[]
    single_hosts=[]
    for net in f:
        if str(net).find('/')!=-1:
            if str(net).find('\n')!=0:
                nets_to_research.append(net[:-1])
            else:
                nets_to_research.append(net)
        else:
            if str(net).find('\n')!=0:
                single_hosts.append(net[:-1])
            else:
                single_hosts.append(net)
    f.close
    if single_hosts!=[]:
        nets_to_research.append(single_hosts)
    return nets_to_research

def getting_shodan_info(path_to_file):
    mass_all_network_info_row=[]
    all_is_ok=0
    for nets in path_to_file:
        mass_network_info_row=[]
        if str(nets).find('/')!=-1:
            print('Net', nets, 'is collecting')
            for ip in ipaddress.ip_network(nets):
                if checkNet:
                    if all_is_ok%64==0:
                        print('Getting info from shodan. We have got info about',all_is_ok,'hosts of current network')
                    all_is_ok+=1
                    try:
                        mass_network_info_row.append(api.host(str(ip)))
                    except:
                        pass
                else:
                    return('There is not internet connection')
            print('Net', nets, 'was made')
        else:
            for ip in nets:
                print('Single_ip', ip, 'was made')
                try:
                    mass_network_info_row.append(api.host(str(ip)))
                except:
                    pass
        log_file = open('all_collected_data.txt', 'a')
        log_file.write(str(mass_network_info_row))
        log_file.close()

        mass_all_network_info_row.append(mass_network_info_row)
        all_is_ok=0

    return mass_all_network_info_row

def get_info_from_row(ip_info_row):
    mass_all_info_about_all_ip_clear=[]

    all_info_about_services=[]
    for services in range(len(ip_info_row['data'])):
        info_about_service=[]
        port=str(ip_info_row['data'][services]['port'])+'/'+str(ip_info_row['data'][services]['transport'])
        service=ip_info_row['data'][services]['_shodan']['module']
        try:
            for cve in ip_info_row['data'][services]['vulns'].keys():
                verified=str(ip_info_row['data'][services]['vulns'][cve]['verified'])+':'
                cvss=str(ip_info_row['data'][services]['vulns'][cve]['cvss'])+'\n'
                references=str(ip_info_row['data'][services]['vulns'][cve]['references'])
                info_about_service.append([cve+':',verified,cvss])
            all_info_about_services.append({port:{service:info_about_service}})
        except:
            info_about_service.append('There is not any vulns')
            all_info_about_services.append({port:{service:info_about_service}})
    
    mass_all_info_about_all_ip_clear={'ip':ip_info_row['ip_str'], 'isp':ip_info_row['isp'],'hostnames':ip_info_row['hostnames'],'os':ip_info_row['os'],'ports':all_info_about_services}
    return mass_all_info_about_all_ip_clear

def write_to_exel(mass_all_info_clear):
    book = openpyxl.Workbook()

    for number_network in range(len(mass_all_info_clear)):  ### for all networks in task
        for network in mass_all_info_clear[number_network].keys(): ### for all ip in network
            sheet_name=network
            if sheet_name!='single hosts':
                sheet_name=sheet_name[:-3]+'_'+sheet_name[len(sheet_name)-2:]

        book.create_sheet(title = sheet_name, index = number_network)
        sheet = book[sheet_name]
        headers=headers_for_exel_list(mass_all_info_clear[number_network])

        for col in range(1, len(headers)+1):    ### putting headers in exel (1st string)
            value = headers[col-1].upper()
            cell = sheet.cell(row = 1, column = col)

            
            cell.value = value
            cell.border = Border(top = Side(border_style='thick', color='FF000000'),right = Side(border_style='thin'), bottom = Side(border_style='thick'), left = Side(border_style='thin'))

            cell.alignment = Alignment(horizontal='center',vertical='center')

        number_IP=0
        help_kof_max=1
        ip_positions=[]
        korr_kof=0
        for row in range(2, len(mass_all_info_clear[number_network][network])+2): ### put info about ports
            ip_positions.append(help_kof_max)
            row+=help_kof_max-1-korr_kof
            for col in range(5, len(headers)+1):     ### putting info about every port (cols 5+)
                value=''
                for number_port_char in range(len(mass_all_info_clear[number_network][network][number_IP]['ports'])):
                    
                    row_changed=row
                    try:
                        value_row=mass_all_info_clear[number_network][network][number_IP]['ports'][number_port_char][headers[col-1]].keys()

                        for value in value_row:
                            ports_info=mass_all_info_clear[number_network][network][number_IP]['ports'][number_port_char][headers[col-1]][value]
                            cell = sheet.cell(row = row, column = col)
                            cell.value = value           
                            cell.fill=PatternFill(fill_type='solid', start_color='00FF00')
                              

                            for port_info in sorted(ports_info, key=lambda item: item[2], reverse=True):
                                row_changed+=1
                                if port_info=='There is not any vulns':
                                    cell = sheet.cell(row = row_changed, column = col)
                                    cell.value = port_info
                                else:
                                    
                                    value=port_info[0]+' '+port_info[1]+' '+port_info[2]
                                    cell = sheet.cell(row = row_changed, column = col)
                                    cell.value = value

                                    scale=float(port_info[2][:-1])
                                    font=''
                                    if scale>=5:
                                        if port_info[1]=='True:' or scale>=7:
                                            cell.fill=PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
                                        if scale>=7 and port_info[1]=='True' or scale>=9:
                                            cell.fill=PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')

                            if help_kof_max<row_changed:
                                help_kof_max=row_changed
                        break
                    except:
                        value='-'
                        cell = sheet.cell(row = row, column = col)
                        cell.value = value
                if value==[] or value==None:
                    value='Unknown'
            number_IP+=1
            korr_kof+=1


        ip_positions.append(help_kof_max)
        number_IP=0
        for row1 in range(len(ip_positions)-1):   ### putting IP, Hostnames, ISP, OS
            row=ip_positions[row1]+1
            for col in range(1, 5):    
                value=mass_all_info_clear[number_network][network][number_IP][headers[col-1]]
                if value==[] or value==None:
                    value='Unknown'
                cell = sheet.cell(row = row, column = col)
                cell.alignment = Alignment(horizontal='center',vertical='center') 
                try:
                    cell.value = value
                except:
                    val_itog=''
                    for val in value:
                        if val_itog!='':
                            val_itog=', '+val_itog
                        val_itog+=val
                    cell.value=val_itog
                
                cell.border = Border(top = Side(border_style='thick', color='FF000000'),right = Side(border_style='thin'), bottom = Side(border_style='thick'), left = Side(border_style='thin'))
                
                sheet.merge_cells(get_column_letter(col)+str(ip_positions[row1]+1)+':'+get_column_letter(col)+str(ip_positions[row1+1]))

            number_IP+=1

        #####     PAINT COL AND ROW LINES     #####
        sheet.column_dimensions['A'].width = 17
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 15
        for col in range(5, len(headers)+1):
            sheet.column_dimensions[get_column_letter(col)].width = 25
        sheet.freeze_panes = 'C2'
        for ip_position in range(1,max(ip_positions)+1):
            for col in range(5,len(headers)+1):
                cell = sheet.cell(row = ip_position, column = col)
                if ip_position in ip_positions:
                    cell.border = Border(right = Side(border_style='thin'), bottom = Side(border_style='thick'), left = Side(border_style='thin'))
                else:
                    cell.border = Border(right = Side(border_style='thin'), left = Side(border_style='thin'))

    book.save('info_about_organization.xlsx')

def headers_for_exel_list(mass_all_info_clear):
    ports_all=[]                    #find all ports for headers
    for network_info_clear in mass_all_info_clear.keys():
        for ip_number in range(len(mass_all_info_clear[network_info_clear])):
            for port in range(len(mass_all_info_clear[network_info_clear][ip_number]['ports'])):
                for port_clear in mass_all_info_clear[network_info_clear][ip_number]['ports'][port].keys():
                    if port_clear not in ports_all:
                        ports_all.append(port_clear)     #end find all ports for headers
    ports_all=sorted(ports_all)     #sort all ports for headers
    helper=[]                       
    ports_all_sorted=[]
    for ports in ports_all:         
        helper.append(int(ports[:ports.find('/')]))
    helper=sorted(helper)
    for ports in helper:
        for ports_help in ports_all:
            if int(ports_help[:ports_help.find('/')])==ports:
                ports_all_sorted.append(ports_help)  #end sort all ports for headers    
    headers_for_exel=[]            #make headers for exel
    headers_for_exel.append('ip')
    headers_for_exel.append('hostnames')
    headers_for_exel.append('isp')
    headers_for_exel.append('os')
    for ports in ports_all_sorted:
        headers_for_exel.append(ports)
    return headers_for_exel       #end make headers for exel

def main():
    # Made by Kraken. Elvis+
    checkNet()
    mass_all_info_clear=[]
    nets_to_research=get_nets_from_file(path_to_file)
    mass_all_network_info_row=getting_shodan_info(nets_to_research)
    nets_to_research=nets_to_research[:-1]
    nets_to_research.append('single hosts')
    i=0

    for network in mass_all_network_info_row:
        mass_ip_info_clear=[]
        for ip_info_row in network:
            if network!=[] and ip_info_row!=[]:
                mass_ip_info_clear.append(get_info_from_row(ip_info_row))
        if mass_ip_info_clear!=[]:
            mass_all_info_clear.append({str(nets_to_research[i]):mass_ip_info_clear})
        i+=1

    write_to_exel(mass_all_info_clear)

main()